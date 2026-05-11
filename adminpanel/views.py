from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Case, When, Value, IntegerField, Q
import openpyxl

from employers.models import Firm, Corporate
from .models import ManageSkill, Domain
from candidates.models import CandidateProfile

from .utils import send_firm_status_email, send_corporate_status_email, send_candidate_status_email


# =========================
# ADMIN DASHBOARD
# =========================
def admin_dashboard(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    context = {
        "total_firms": Firm.objects.count(),
        "pending_firms": Firm.objects.filter(status='pending').count(),
        "approved_firms": Firm.objects.filter(status='approved').count(),
        "rejected_firms": Firm.objects.filter(status='rejected').count(),

        "total_corporates": Corporate.objects.count(),
        "pending_corporates": Corporate.objects.filter(status='pending').count(),
        "approved_corporates": Corporate.objects.filter(status='approved').count(),
        "rejected_corporates": Corporate.objects.filter(status='rejected').count(),

        "total_candidates": CandidateProfile.objects.count(),
        "pending_candidates": CandidateProfile.objects.filter(status='pending').count(),
        "approved_candidates": CandidateProfile.objects.filter(status='approved').count(),
        "rejected_candidates": CandidateProfile.objects.filter(status='rejected').count(),
    }
    return render(request, "adminpanel/admin_dashboard.html", context)


# =========================
# VERIFY EMPLOYERS
# =========================
def verify_employers(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')
    return render(request, 'adminpanel/verify_employers.html')


# =========================
# FIRMS VERIFICATION
# =========================
def verify_firm(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    firms = Firm.objects.annotate(
        status_order=Case(
            When(status='pending', then=Value(0)),
            When(status='approved', then=Value(1)),
            When(status='rejected', then=Value(2)),
            output_field=IntegerField()
        )
    ).order_by('status_order', '-created_at')

    context = {
        "firms": firms,
        "total_firms": Firm.objects.count(),
        "pending_firms": Firm.objects.filter(status='pending').count(),
        "approved_firms": Firm.objects.filter(status='approved').count(),
        "rejected_firms": Firm.objects.filter(status='rejected').count(),
    }
    return render(request, "adminpanel/verify_firm.html", context)


def update_firm_status(request, firm_id, action):
    if not request.session.get('admin_login'):
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=401)

    firm = get_object_or_404(Firm, id=firm_id)

    if action == "approve":
        firm.status = "approved"
        msg = f"{firm.name} approved successfully"
    elif action == "reject":
        firm.status = "rejected"
        msg = f"{firm.name} rejected successfully"
    else:
        return JsonResponse({'success': False, 'error': 'Invalid action'}, status=400)

    firm.save()
    send_firm_status_email(firm, action)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'new_status': firm.status, 'message': msg})

    messages.success(request, msg)
    return redirect('adminpanel:verify_firm')


def view_firm_detail(request, firm_id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    firm = get_object_or_404(Firm, id=firm_id)
    return render(request, 'adminpanel/firm_detail.html', {'firm': firm})


# =========================
# CORPORATES VERIFICATION
# =========================
def verify_corporate(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    corporates = Corporate.objects.annotate(
        status_order=Case(
            When(status='pending', then=Value(0)),
            When(status='approved', then=Value(1)),
            When(status='rejected', then=Value(2)),
            output_field=IntegerField()
        )
    ).order_by('status_order', '-id')

    context = {
        "corporates": corporates,
        "total_corporates": Corporate.objects.count(),
        "pending_corporates": Corporate.objects.filter(status='pending').count(),
        "approved_corporates": Corporate.objects.filter(status='approved').count(),
        "rejected_corporates": Corporate.objects.filter(status='rejected').count(),
    }
    return render(request, "adminpanel/verify_corporate.html", context)


def update_corporate_status(request, corporate_id, action):
    if not request.session.get('admin_login'):
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=401)

    corporate = get_object_or_404(Corporate, id=corporate_id)

    if action == "approve":
        corporate.status = "approved"
        corporate.is_verified = True
        msg = f"{corporate.name} approved successfully"
    elif action == "reject":
        corporate.status = "rejected"
        corporate.is_verified = False
        msg = f"{corporate.name} rejected successfully"
    else:
        return JsonResponse({'success': False, 'error': 'Invalid action'}, status=400)

    corporate.save()
    send_corporate_status_email(corporate, action)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'new_status': corporate.status, 'message': msg})

    messages.success(request, msg)
    return redirect('adminpanel:verify_corporate')


def view_corporate_detail(request, corporate_id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    corporate = get_object_or_404(Corporate, id=corporate_id)
    return render(request, 'adminpanel/corporate_detail.html', {'corporate': corporate})


# =========================
# CANDIDATES VERIFICATION
# =========================
def verify_candidate(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    candidates = CandidateProfile.objects.select_related('user').annotate(
        status_order=Case(
            When(status='pending', then=Value(0)),
            When(status='approved', then=Value(1)),
            When(status='rejected', then=Value(2)),
            output_field=IntegerField()
        )
    ).order_by('status_order', '-created_at')

    context = {
        "candidates": candidates,
        "total_candidates": CandidateProfile.objects.count(),
        "pending_candidates": CandidateProfile.objects.filter(status='pending').count(),
        "approved_candidates": CandidateProfile.objects.filter(status='approved').count(),
        "rejected_candidates": CandidateProfile.objects.filter(status='rejected').count(),
    }
    return render(request, "adminpanel/verify_candidate.html", context)


def update_candidate_status(request, candidate_id, action):
    if not request.session.get('admin_login'):
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=401)

    candidate = get_object_or_404(CandidateProfile, id=candidate_id)

    if action == "approve":
        candidate.status = "approved"
        candidate.is_verified = True
        msg = "Candidate approved successfully"
    elif action == "reject":
        candidate.status = "rejected"
        candidate.is_verified = False
        msg = "Candidate rejected successfully"
    else:
        return JsonResponse({'success': False, 'error': 'Invalid action'}, status=400)

    candidate.save()
    send_candidate_status_email(candidate, action)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({"success": True, "new_status": candidate.status, "message": msg})

    messages.success(request, msg)
    return redirect('adminpanel:verify_candidates')


def view_candidate_detail(request, candidate_id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    candidate = get_object_or_404(CandidateProfile, id=candidate_id)
    return render(request, 'adminpanel/candidate_detail.html', {'candidate': candidate})


# =========================
# SKILLS & DOMAINS CRUD
# =========================
def skill_list(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    query = request.GET.get('q', '')
    if query:
        skills = ManageSkill.objects.filter(
            Q(skill_name__icontains=query) |
            Q(category__icontains=query)
        ).order_by('-id')
    else:
        skills = ManageSkill.objects.all().order_by('-id')

    domains = Domain.objects.all().order_by('-id')
    return render(request, 'manage_skill/skill_list.html', {'skills': skills, 'domains': domains, 'query': query})

#
# def skill_add(request):
#     if not request.session.get('admin_login'):
#         messages.error(request, "Admin access required. Please log in.")
#         return redirect('login')
#
#     if request.method == "POST":
#         ManageSkill.objects.create(
#             skill_name=request.POST.get('skill_name'),
#             category=request.POST.get('category')
#         )
#         return redirect('adminpanel:manage_skills')
#
#     context = {}
#     if hasattr(ManageSkill, 'CATEGORY_CHOICES'):
#         context['categories'] = ManageSkill.CATEGORY_CHOICES
#     return render(request, 'manage_skill/skill_form.html', context)

def skill_add(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    if request.method == "POST":
        ManageSkill.objects.create(
            skill_name=request.POST.get('skill_name'),
            category=request.POST.get('category'),
            proficiency=request.POST.get('proficiency')  # Save the new field
        )
        return redirect('adminpanel:manage_skills')

    context = {}
    if hasattr(ManageSkill, 'CATEGORY_CHOICES'):
        context['categories'] = ManageSkill.CATEGORY_CHOICES

    if hasattr(ManageSkill, 'PROFICIENCY_CHOICES'):
        context['proficiencies'] = ManageSkill.PROFICIENCY_CHOICES

    return render(request, 'manage_skill/skill_form.html', context)

def skill_edit(request, id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    skill = get_object_or_404(ManageSkill, id=id)
    if request.method == "POST":
        skill.skill_name = request.POST.get('skill_name')
        skill.category = request.POST.get('category')
        skill.proficiency = request.POST.get('proficiency') # Save changes
        skill.save()
        return redirect('adminpanel:manage_skills')

    context = {
        'skill': skill,
        'categories': ManageSkill.CATEGORY_CHOICES,
        'proficiencies': ManageSkill.PROFICIENCY_CHOICES
    }
    return render(request, 'manage_skill/skill_form.html', context)

#
# def skill_edit(request, id):
#     if not request.session.get('admin_login'):
#         messages.error(request, "Admin access required. Please log in.")
#         return redirect('login')
#
#     skill = get_object_or_404(ManageSkill, id=id)
#     if request.method == "POST":
#         skill.skill_name = request.POST.get('skill_name')
#         skill.category = request.POST.get('category')
#         skill.save()
#         return redirect('adminpanel:manage_skills')
#
#     context = {'skill': skill}
#     if hasattr(ManageSkill, 'CATEGORY_CHOICES'):
#         context['categories'] = ManageSkill.CATEGORY_CHOICES
#     return render(request, 'manage_skill/skill_form.html', context)
#

def skill_delete(request, id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    skill = get_object_or_404(ManageSkill, id=id)
    if request.method == "POST":
        skill.delete()
        return redirect('adminpanel:manage_skills')
    return render(request, 'manage_skill/skill_delete.html', {'skill': skill})


def domain_add(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    if request.method == "POST":
        name = request.POST.get('domain_name').strip()
        if name:
            Domain.objects.get_or_create(domain_name=name)
        return redirect('adminpanel:domain_add')
    domains = Domain.objects.all().order_by('-id')
    return render(request, 'manage_skill/domain_add.html', {'domains': domains})


def domain_edit(request, id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    domain = get_object_or_404(Domain, id=id)
    if request.method == "POST":
        new_name = request.POST.get("domain_name")
        if new_name:
            domain.domain_name = new_name
            domain.save()
        return redirect('adminpanel:domain_add')
    domains = Domain.objects.all()
    return render(request, 'manage_skill/domain_edit.html', {'domain_to_edit': domain, 'domains': domains})


def domain_delete(request, id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    domain = get_object_or_404(Domain, id=id)
    if request.method == "POST":
        domain.delete()
        return redirect('adminpanel:domain_add')
    return render(request, 'manage_skill/domain_delete.html', {'domain': domain})


# =========================
# EXPORT EXCEL FUNCTIONS (FIXED)
# =========================

def export_skills_excel(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(['ID', 'Skill Name', 'Category', 'Created At'])

    for skill in ManageSkill.objects.all().order_by('-id'):
        ws.append([
            skill.id,
            skill.skill_name,
            getattr(skill, 'category', ''),
            skill.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(skill, 'created_at') else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=skills.xlsx'
    wb.save(response)
    return response


def export_single_firm_excel(request, firm_id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    try:
        firm = Firm.objects.get(id=firm_id)
    except Firm.DoesNotExist:
        return HttpResponse("Firm not found", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firm Details"

    headers = [
        "Firm Name", "ICAI Registration No", "City / Area", "No. of Partners",
        "No. of Paid Assistants", "Articleship Seats", "Jobs Available",
        "Specialization", "Exposure Level", "Work Hours", "Stipend Range",
        "Leave Policy", "Mentorship Available", "Status"
    ]
    ws.append(headers)

    field_map = {
        "Firm Name": "name", "ICAI Registration No": "registration_number", "City / Area": "city_area",
        "No. of Partners": "no_of_partners", "No. of Paid Assistants": "no_of_paid_assistants",
        "Articleship Seats": "articleship_seats", "Jobs Available": "jobs_available",
        "Specialization": "specialization", "Exposure Level": "exposure_level", "Work Hours": "work_hours",
        "Stipend Range": "stipend_range", "Leave Policy": "leave_policy",
        "Mentorship Available": "mentorship_available", "Status": "status"
    }

    row = []
    for header in headers:
        field_name = field_map.get(header)
        value = getattr(firm, field_name, "")
        if isinstance(value, bool):
            value = "Yes" if value else "No"
        row.append(str(value) if value is not None else "")
    ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=firm_{firm_id}.xlsx'
    wb.save(response)
    return response


def export_all_firms_excel(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    firms = Firm.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Firms"

    headers = [
        "Firm Name", "ICAI Registration No", "City / Area", "No. of Partners",
        "No. of Paid Assistants", "Articleship Seats", "Jobs Available",
        "Specialization", "Exposure Level", "Work Hours", "Stipend Range",
        "Leave Policy", "Mentorship Available", "Status"
    ]
    ws.append(headers)

    field_map = {
        "Firm Name": "name", "ICAI Registration No": "registration_number", "City / Area": "city_area",
        "No. of Partners": "no_of_partners", "No. of Paid Assistants": "no_of_paid_assistants",
        "Articleship Seats": "articleship_seats", "Jobs Available": "jobs_available",
        "Specialization": "specialization", "Exposure Level": "exposure_level", "Work Hours": "work_hours",
        "Stipend Range": "stipend_range", "Leave Policy": "leave_policy",
        "Mentorship Available": "mentorship_available", "Status": "status"
    }

    for firm in firms:
        row = []
        for header in headers:
            value = getattr(firm, field_map.get(header), "")
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            row.append(str(value) if value is not None else "")
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=all_firms.xlsx'
    wb.save(response)
    return response


def export_single_corporate_excel(request, corporate_id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    try:
        corp = Corporate.objects.get(id=corporate_id)
    except Corporate.DoesNotExist:
        return HttpResponse("Corporate not found", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corporate Details"

    headers = [
        "Name", "Email", "Mobile", "ICAI Reg Number", "City/Area",
        "Finance Exposure", "Hiring Type", "Work Model",
        "Jobs Available", "About", "CA Hiring", "Is Verified", "Status"
    ]
    ws.append(headers)

    field_map = {
        "Name": "name", "Email": "email", "Mobile": "mobile", "ICAI Reg Number": "registration_number",
        "City/Area": "city_area", "Finance Exposure": "finance_exposure",
        "Hiring Type": "hiring_type", "Work Model": "work_model", "Jobs Available": "jobs_available",
        "About": "about", "CA Hiring": "ca_hiring", "Is Verified": "is_verified", "Status": "status"
    }

    row = []
    for header in headers:
        value = getattr(corp, field_map[header], "")
        if isinstance(value, bool):
            value = "Yes" if value else "No"
        if isinstance(value, list) or isinstance(value, dict):
            value = ", ".join([str(v) for v in value])
        row.append(str(value))
    ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=corporate_{corporate_id}.xlsx'
    wb.save(response)
    return response


def export_all_corporates_excel(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    corporates = Corporate.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Corporates"

    headers = [
        "Name", "Email", "Mobile", "ICAI Reg Number", "City/Area",
        "Finance Exposure", "Hiring Type", "Work Model",
        "Jobs Available", "About", "CA Hiring", "Is Verified", "Status"
    ]
    ws.append(headers)

    field_map = {
        "Name": "name", "Email": "email", "Mobile": "mobile", "ICAI Reg Number": "registration_number",
        "City/Area": "city_area", "Finance Exposure": "finance_exposure",
        "Hiring Type": "hiring_type", "Work Model": "work_model", "Jobs Available": "jobs_available",
        "About": "about", "CA Hiring": "ca_hiring", "Is Verified": "is_verified", "Status": "status"
    }

    for corp in corporates:
        row = []
        for header in headers:
            value = getattr(corp, field_map[header], "")
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            if isinstance(value, list) or isinstance(value, dict):
                value = ", ".join([str(v) for v in value])
            row.append(str(value))
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=all_corporates.xlsx'
    wb.save(response)
    return response


def export_single_candidate_excel(request, candidate_id):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    try:
        candidate = CandidateProfile.objects.get(id=candidate_id)
    except CandidateProfile.DoesNotExist:
        return HttpResponse("Candidate not found", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Details"

    headers = [
        "Email", "CA Status", "ICAI Reg Number", "Languages Known", "Availability",
        "Foundation Year", "Foundation Attempts", "Intermediate Attempts G1", "Intermediate Attempts G2",
        "Preferred Articleship City", "Industry Preference", "Preferred Job Roles", "Employment Type",
        "Final Group Appeared", "Final Attempts", "Career Preference", "Study Status",
        "Software Skills", "Articleship Start Date", "Articleship End Date",
        "Current Firm City", "Is Confidential Mode", "Is Verified", "Status"
    ]
    ws.append(headers)

    field_map = {
        "Email": "user.email", "CA Status": "user.ca_status", "ICAI Reg Number": "icai_reg_num",
        "Languages Known": "languages", "Availability": "availability", "Foundation Year": "foundation_year",
        "Foundation Attempts": "foundation_attempts", "Intermediate Attempts G1": "inter_attempts_g1",
        "Intermediate Attempts G2": "inter_attempts_g2", "Preferred Articleship City": "preferred_articleship_city",
        "Industry Preference": "industry_preference", "Preferred Job Roles": "preferred_job_roles",
        "Employment Type": "employment_type", "Final Group Appeared": "final_group_appeared",
        "Final Attempts": "final_attempts", "Career Preference": "career_preference",
        "Study Status": "study_status", "Software Skills": "software_skills",
        "Articleship Start Date": "articleship_start_date", "Articleship End Date": "articleship_end_date",
        "Current Firm City": "current_firm_city", "Is Confidential Mode": "is_confidential_mode",
        "Is Verified": "is_verified", "Status": "status"
    }

    row = []
    for header in headers:
        field_name = field_map.get(header)
        value = candidate
        for attr in field_name.split("."):
            value = getattr(value, attr, "")
            if value is None: value = ""

        if isinstance(value, bool): value = "Yes" if value else "No"
        if hasattr(value, "all"): value = ", ".join([str(s) for s in value.all()])
        if isinstance(value, list) or isinstance(value, dict): value = str(value)
        row.append(str(value))

    ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=candidate_{candidate_id}.xlsx'
    wb.save(response)
    return response


def export_all_candidates_excel(request):
    if not request.session.get('admin_login'):
        messages.error(request, "Admin access required. Please log in.")
        return redirect('login')

    candidates = CandidateProfile.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Candidates"

    headers = [
        "Email", "CA Status", "ICAI Reg Number", "Languages Known", "Availability",
        "Foundation Year", "Foundation Attempts", "Intermediate Attempts G1", "Intermediate Attempts G2",
        "Preferred Articleship City", "Industry Preference", "Preferred Job Roles", "Employment Type",
        "Final Group Appeared", "Final Attempts", "Career Preference", "Study Status",
        "Software Skills", "Articleship Start Date", "Articleship End Date",
        "Current Firm City", "Is Confidential Mode", "Is Verified", "Status"
    ]
    ws.append(headers)

    field_map = {
        "Email": "user.email", "CA Status": "user.ca_status", "ICAI Reg Number": "icai_reg_num",
        "Languages Known": "languages", "Availability": "availability", "Foundation Year": "foundation_year",
        "Foundation Attempts": "foundation_attempts", "Intermediate Attempts G1": "inter_attempts_g1",
        "Intermediate Attempts G2": "inter_attempts_g2", "Preferred Articleship City": "preferred_articleship_city",
        "Industry Preference": "industry_preference", "Preferred Job Roles": "preferred_job_roles",
        "Employment Type": "employment_type", "Final Group Appeared": "final_group_appeared",
        "Final Attempts": "final_attempts", "Career Preference": "career_preference",
        "Study Status": "study_status", "Software Skills": "software_skills",
        "Articleship Start Date": "articleship_start_date", "Articleship End Date": "articleship_end_date",
        "Current Firm City": "current_firm_city", "Is Confidential Mode": "is_confidential_mode",
        "Is Verified": "is_verified", "Status": "status"
    }

    for candidate in candidates:
        row = []
        for header in headers:
            field_name = field_map.get(header)
            value = candidate
            for attr in field_name.split("."):
                value = getattr(value, attr, "")
                if value is None: value = ""

            if isinstance(value, bool): value = "Yes" if value else "No"
            if hasattr(value, "all"): value = ", ".join([str(s) for s in value.all()])
            if isinstance(value, list) or isinstance(value, dict): value = str(value)
            row.append(str(value))
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=all_candidates.xlsx'
    wb.save(response)
    return response